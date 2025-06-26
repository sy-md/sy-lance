import pandas as pd
import openpyxl as op

# find the location fo the data
data = "/mnt/c/Users/crazy/Downloads/W PHX INVENTORY.xlsx"
data1 = "W PHX INVENTORY_updated.xlsx"
#read the data
my_file = pd.read_excel(data)
#print(my_file.head())

#activate the data
wb = op.load_workbook(data)
sheet = wb.active

#manipulate the data
rows = sheet.max_row
col = 4

polaris = {
        "h050" : 120,
        "b012" : 394,
        "h224" : 288,
        "h089" : 322,
        "ekp02" : 770,
        "etk02-pl" : 736,
        "ecm01-can" : 732
        }

new_cols = sheet["E3"]

print("before" + "{}".format(sheet["E3"].value))
for i in range(1, 4): #starting in the first row till max row
    for j in range(1, col + 1): # 
        cell_obj = sheet.cell(row=i, column=j)
        if cell_obj.value == "H294":
            print("found")
            print(i,j)
            sheet["E3"] = "test"
        print(cell_obj.value, end= " - ")
    print() # Add a newline at the end of the row
       

print(sheet["E3"].value)


wb.save(data1)
