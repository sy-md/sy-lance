import pandas as pd
import openpyxl as op

# find the location fo the data
data1 = "/mnt/c/Users/crazy/Downloads/W PHX INVENTORY.xlsx"
#data1 = "W PHX INVENTORY_updated.xlsx"
#read the data
my_file = pd.read_excel(data1)
#print(my_file.head())

#activate the data
wb = op.load_workbook(data1)
sheet = wb.active

#manipulate the data
col = 4
#rows = sheet.max_row
rows = 10

"""
i need to rethink this so here is the breakdown as of 6/27/25 12:14

10 polaris
1 tqkeuchi [050,067,ecm,mc4,b012,ekpo2] 

kelly: uncounted for
    1 keypad 


find the sims in the vechils every vevhical use the same stuff in some deg
"""

# create a dumper of part being a dump_parts.py
parts = {
        "H050" : [1000120 , 0 ], 
        "B012" : [1000394, 0 ],
        "H244" : [1000288, 0],
        "H089" :[1000322, 0],
        "EKP02" : [1000770, 0],
        "ETK02-2PL" :  [1000736, 0],
        "ECM01-CAN" : [1000732, 0], 
        "ETK12-SNT" : [10006041, 0] #slap n track
        }

#move this in a vehcils.json  
"""
                instead of putting them here and re-typing the interface
                will do that for i just did a [polaris] and used 1 of this
                that and those and ill populate this for you 

                also make a make ready so it know all the piece | youre going
                to use all ready for exmaple keep this under   V

"""

vechils = {
        "polaris" : [
            parts["H050"][0],
            parts["B012"][0],
            parts["H244"][0],
            parts["H089"][0],
            parts["EKP02"][0],
            parts["ETK02-2PL"][0],
            parts["ECM01-CAN"][0]
        ],
        "Kubbota": [],
        "case" : [],
        "takeuchi" : []
}


"""
data = {
    inventory : {"name" : id}
    vehcial : []
}


"""

# todo list:

# crud operations
# tmp subbing of part
#interface for quick operations


# a cli for how man unit you did



lst = []
for i in range(1, rows): #starting in the first row till max row
    for j in range(1, col + 1): # jump colums till max
        cell_obj = sheet.cell(row=i, column=j) # storing curr location

        if cell_obj.value in vechils["polaris"]:
            print("searching for {}".format(cell_obj.value))
            #oh = sheet.cell(row=i,column=j+3) # after finding the name just to on hand
            #sheet["E{}".format(i)] = oh.value - units # subtract units done from on hand

        print(cell_obj.value, end= " - ")
    print() # Add a newline at the end of the row


       

print(sheet["E3"].value)

#nothing

#wb.save(data1)
