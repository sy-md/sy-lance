import pandas as pd
import openpyxl as op
import json as js

# find the location fo the data
#data1 = "/mnt/c/Users/crazy/Downloads/W PHX INVENTORY.xlsx"
#data1 = "/mnt/c/Users/crazy/Downloads/W PHX INVENTORY.xlsx"
data1 = "W PHX INVENTORY_updated.xlsx"
#read the data
my_file = pd.read_excel(data1)
#print(my_file.head())

#activate the data
wb = op.load_workbook(data1)
sh = wb.active

#manipulate the data
col = 4
#rows = sheet.max_row
rows = 10

"""
i need to rethink this so here is the breakdown as of 6/27/25 12:14

10 polaris
1 tqkeuchi [050,067,ecm,mc4,b012,ekpo2] 

kelly: uncounted for
    1 keypad 


find the sims in the vechils every vevhical use the same stuff in some deg
"""
"""
                instead of putting them here and re-typing the interface
                will do that for i just did a [polaris] and used 1 of this
                that and those and ill populate this for you 

                also make a make ready so it know all the piece | youre going
                to use all ready for exmaple keep this under   V

"""
parts = {
    "H050" : [1000120], 
    "B012" : [1000394],
    "H244" : [1000288],
    "H089" :[1000322],
    "EKP02" : [1000770],
    "ETK02-2PL" :  [1000736],
    "ECM01-CAN" : [1000732], 
    "ETK12-SNT" : [10006041] #slap n track
}

data = {
    "polaris" : {
        "cable" : parts["H050"],
        "six" :parts["B012"],
        "five" :parts["H244"],
        "four" :parts["H089"],
        "three" :parts["EKP02"],
        "two" :parts["ETK02-2PL"],
        "one" :parts["ECM01-CAN"]
    },
    "vechils" : {},
    "kubbota" : {},
    "case" : {},
    "takeuchi" : {},
}

# crud operations
# tmp subbing of part
#interface for quick operations
amt = 54

# a cli for how man unit you did

def interface():
    x = input("how how many of the X of units have you installed - ") # this will be the promts
    the_type = input("i need the show a list here (the type of vechil) - ")
    
   

    if the_type in data.keys(): # the unit you just worked on is the the type
        lst = [] # put the type of parts in this list


        for v in data[the_type].values():
            lst.append(v[0])
            print("the list of parts of the unit you are looking for")

        with open("data.json", "r") as file:
            my_data = js.load(file)
            for i in range(1, rows):
                for j in range(1, col+ 1):
                    cell = sh.cell(row=i, column=j) # gets the cell info from the cell
                
                    if cell.value in lst:
                        for d in my_data[the_type].values():
                            if len(d) == 2:
                                print(len(d))
                                d[1] += int(x)
                            else:
                                d.append(int(x))
                                print("added to the database")

            with open("data.json", "w") as file:
                js.dump(my_data, file, indent=4)

            

            oh = sh.cell(row=i,column=j+3) # after finding the name just to on hand
            sh["E{}".format(i)] = oh.value - x # subtract units done from on hand



            



#   if the_type in data.keys():
#       lst = []
#       for i in range(1, rows): #starting in the first row till max row
#           for j in range(1, col + 1): # jump colums till max
#               cell_obj = sheet.cell(row=i, column=j) # storing curr location

#               print("choice was {}".format(the_type))

#               5
#               
#               if cell_obj.value == data[the_type][0+1]:
#                   print("is in the data dict")
#                   for k,v in data[the_type].items():
#                       print("the value is"+str(v))
#                       print("searching for {} with {}".format(cell_obj.value,x))
#                   #oh = sheet.cell(row=i,column=j+3) # after finding the name just to on hand
#                   #sheet["E{}".format(i)] = oh.value - units # subtract units done from on hand
#               print("not found")
#               print(cell_obj.value, end= " - ")
#           print() # Add a newline at the end of the row
#       print(sheet["E3"].value)


#       if the_type not in data:
#           data[the_type] = {}
#           print("added this to the dict {}".format(data) )










interface()
wb.save(data1)
